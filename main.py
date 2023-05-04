from flask import Flask, render_template,jsonify,request
from openpyxl import load_workbook
import pandas as pd

app = Flask(__name__)

@app.route('/')
def display_table():
    # Load the Excel file
    wb = load_workbook('BBDD.xlsx')
    ws = wb.active

    # Get the data from the Excel file
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Render the template with the data
    return render_template('table.html', data=data)

@app.route('/data', methods = ['GET'])
def give_table():
    if(request.method == 'GET'):
        a = pd.read_excel("BBDD.xlsx")
        a = a.to_dict()
        return jsonify(a)

app.run(host='0.0.0.0', port=81)
